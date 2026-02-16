from __future__ import annotations

import logging
from dataclasses import dataclass
from typing import Callable, Optional, Tuple

import pandas as pd
from openpyxl.utils import get_column_letter

from utils import slugify_columns
from report import build_report

logger = logging.getLogger(__name__)
ProgressFn = Callable[[float, str], None]  # (0.0..1.0, mensagem)


@dataclass
class ProcessResult:
    out_xlsx: str
    out_report: str
    linhas_before: int
    linhas_after: int
    erros_datas: int
    erros_valores: int


def _read_file(path: str) -> pd.DataFrame:
    lower = path.lower()
    if lower.endswith(".csv"):
        try:
            return pd.read_csv(path, sep=",", encoding="utf-8")
        except Exception:
            return pd.read_csv(path, sep=";", encoding="utf-8")
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        return pd.read_excel(path)
    raise ValueError("Formato não suportado. Use CSV ou XLSX.")


def _to_float_money(series: pd.Series) -> Tuple[pd.Series, pd.Series]:
    """
    Converte textos do tipo 'R$ 1.234,56' ou '99,9' em float.
    Retorna (serie_float, mask_erro) onde mask_erro=True significa 'valor inválido' (não vazio).
    """
    s = series.astype("string").str.strip()

    # considera vazio como vazio (não erro)
    vazio = s.isna() | (s == "") | (s.str.lower() == "nan")

    s = s.str.replace("R$", "", regex=False).str.strip()
    s = s.str.replace(".", "", regex=False).str.replace(",", ".", regex=False)

    nums = pd.to_numeric(s, errors="coerce")
    erro = (~vazio) & (nums.isna())
    return nums, erro


def _to_date(series: pd.Series) -> Tuple[pd.Series, pd.Series]:
    """
    Converte strings de datas variadas em datetime.
    Retorna (serie_datetime, mask_erro) onde mask_erro=True significa 'data inválida' (não vazia).
    """
    s = series.astype("string").str.strip()
    vazio = s.isna() | (s == "") | (s.str.lower() == "nan")

    dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
    erro = (~vazio) & (dt.isna())
    return dt, erro


def _basic_clean_and_errors(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """
    Retorna:
    - df_limpo (para salvar)
    - df_erros (linhas com problemas)
    - stats (contagens)
    """
    out = df.copy()
    out.columns = slugify_columns(list(out.columns))

    # Normalização básica de strings
    for col in out.columns:
        if out[col].dtype == "object":
            out[col] = out[col].astype("string").str.strip()
            out[col] = out[col].str.replace(r"\s+", " ", regex=True)

    # Observação: preencher vazio
    if "observao" in out.columns:  # no seu exemplo virou 'observao'
        out["observao"] = out["observao"].astype("string").str.strip()
        out["observao"] = out["observao"].replace("", pd.NA).fillna("Não informado")

    # Nome: title case
    if "nome" in out.columns:
        out["nome"] = out["nome"].astype("string").str.strip().str.title()

    # Email: lowercase
    if "email" in out.columns:
        out["email"] = out["email"].astype("string").str.strip().str.lower()

    # Valor: converter
    erros_valores = pd.Series([False] * len(out))
    if "valor_r" in out.columns:
        out["valor_r"], erros_valores = _to_float_money(out["valor_r"])

    # Data: converter (mantém datetime para formatar no Excel)
    erros_datas = pd.Series([False] * len(out))
    if "data_de_pagamento" in out.columns:
        out["data_de_pagamento"], erros_datas = _to_date(out["data_de_pagamento"])

    # Remover duplicadas (depois das normalizações)
    out = out.drop_duplicates()

    # Criar aba de erros (pega as linhas com erro, com coluna indicando o motivo)
    # Nota: após drop_duplicates, índices mudam; então recalculamos máscaras alinhadas
    # para a versão "out" usando merge por todas as colunas originais fica pesado.
    # Estratégia simples e confiável: gerar erros ANTES do drop_duplicates e montar df_erros dali.
    # Para manter simples: montamos erros da cópia original *antes* do drop_duplicates.
    base = df.copy()
    base.columns = slugify_columns(list(base.columns))
    for col in base.columns:
        if base[col].dtype == "object":
            base[col] = base[col].astype("string").str.strip().str.replace(r"\s+", " ", regex=True)

    motivos = []
    mask_any = pd.Series([False] * len(base))

    if "valor_r" in base.columns:
        _, mval = _to_float_money(base["valor_r"])
        motivos.append(("valor_invalido", mval))
        mask_any |= mval

    if "data_de_pagamento" in base.columns:
        _, mdt = _to_date(base["data_de_pagamento"])
        motivos.append(("data_invalida", mdt))
        mask_any |= mdt

    df_erros = base[mask_any].copy()
    if len(df_erros) > 0:
        # Coluna "erros" com motivos concatenados
        err_col = []
        for i in df_erros.index:
            ms = []
            for nome, m in motivos:
                if bool(m.loc[i]):
                    ms.append(nome)
            err_col.append(", ".join(ms))
        df_erros["erros"] = err_col

        # Para “datas não informadas”: você pediu ou "Não informado" ou mandar pra erros.
        # Aqui erros = inválidas; ausentes ficam como "Não informado" na saída.
    stats = {
        "erros_valores": int(erros_valores.sum()) if isinstance(erros_valores, pd.Series) else 0,
        "erros_datas": int(erros_datas.sum()) if isinstance(erros_datas, pd.Series) else 0,
    }
    return out, df_erros, stats


def _write_excel_with_formats(df: pd.DataFrame, df_erros: pd.DataFrame, out_xlsx: str):
    """
    Salva df em Excel com formatos:
    - data_de_pagamento -> dd/mm/aaaa
    - valor_r -> moeda R$
    E cria aba ERROS se houver.
    """
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="DADOS")
        if df_erros is not None and len(df_erros) > 0:
            df_erros.to_excel(writer, index=False, sheet_name="ERROS")

        wb = writer.book
        ws = wb["DADOS"]

        # Mapear colunas pelo nome
        col_map = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}

        # Formato moeda
        if "valor_r" in col_map:
            c = col_map["valor_r"]
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=c).number_format = 'R$ #,##0.00'

        # Formato data
        if "data_de_pagamento" in col_map:
            c = col_map["data_de_pagamento"]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=c)
                # Se a célula tiver datetime, aplica formato. Se tiver None, deixa vazio (vamos preencher depois).
                cell.number_format = "DD/MM/YYYY"

        # Autosize simples (opcional)
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = max(12, min(35, ws.column_dimensions[letter].width or 12))


def process_file(
    input_path: str,
    out_xlsx: str,
    out_report: str,
    progress: Optional[ProgressFn] = None,
) -> ProcessResult:
    def p(val: float, msg: str):
        if progress:
            progress(val, msg)

    p(0.05, "Lendo arquivo…")
    df_before = _read_file(input_path)

    p(0.35, "Limpando e padronizando…")
    df_after, df_erros, stats = _basic_clean_and_errors(df_before)

    # Datas ausentes -> "Não informado" (como texto) PARA VISUAL; mas cuidado:
    # Se você quiser manter coluna como data de verdade, a célula vazia é melhor.
    # Você pediu "Não informado", então vamos transformar só os vazios para texto.
    if "data_de_pagamento" in df_after.columns:
        # dt -> string dd/mm/aaaa, vazios -> "Não informado"
        dt = df_after["data_de_pagamento"]
        df_after["data_de_pagamento"] = (
             df_after["data_de_pagamento"]
            .dt.strftime("%d/%m/%Y")
            )

        df_after["data_de_pagamento"] = df_after["data_de_pagamento"].fillna("Não informado")

    p(0.70, "Gerando relatório…")
    report_text = build_report(df_before, df_after, df_erros, stats)

    p(0.85, "Salvando Excel…")
    _write_excel_with_formats(df_after, df_erros, out_xlsx)

    p(0.95, "Salvando relatório…")
    with open(out_report, "w", encoding="utf-8") as f:
        f.write(report_text)

    p(1.0, "Concluído ✅")

    return ProcessResult(
        out_xlsx=out_xlsx,
        out_report=out_report,
        linhas_before=len(df_before),
        linhas_after=len(df_after),
        erros_datas=int(stats.get("erros_datas", 0)),
        erros_valores=int(stats.get("erros_valores", 0)),
    )
